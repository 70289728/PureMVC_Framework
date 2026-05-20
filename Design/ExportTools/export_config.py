"""
Config Table Export Tool
Excel (.xlsx) -> JSON / Lua / C# Class
Supports: int, float, string, bool, list<T>, map<K,V>
"""

import os
import sys
import json
import re
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================
# Paths
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PUREMVC_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "PureMVC_Framework"))
SERVER_ROOT = os.path.join(PUREMVC_ROOT, "ProtoServer")
DESIGN_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))
EXCEL_DIR = os.path.join(DESIGN_DIR, "Excel")
EXPORT_TOOLS_DIR = os.path.join(DESIGN_DIR, "ExportTools")

CLIENT_PROJECT = PUREMVC_ROOT
SERVER_PROJECT = os.path.join(SERVER_ROOT, "ProtoServer")

CLIENT_JSON_DIR = os.path.join(CLIENT_PROJECT, "Assets", "GameConfig")
CLIENT_LUA_DIR = os.path.join(CLIENT_PROJECT, "Assets", "GameConfig", "Lua")
CLIENT_CS_DIR = os.path.join(CLIENT_PROJECT, "Assets", "Scripts", "HotUpdateAssembly", "GameConfigCs")
SERVER_CS_DIR = os.path.join(SERVER_PROJECT, "Config")

# ============================================================
# Type Parsing
# ============================================================

TYPE_PATTERN = re.compile(
    r"^(?P<base>int|float|string|bool)"
    r"(?:"
    r"<(?P<sub1>int|float|string|bool),(?P<sub2>int|float|string|bool)>"
    r")?$"
)

LIST_TYPE_PATTERN = re.compile(r"^list<(?P<elem>int|float|string|bool)>$")
MAP_TYPE_PATTERN = re.compile(
    r"^map<(?P<key>int|float|string|bool),(?P<val>int|float|string|bool)>$"
)

CSHARP_TYPE_MAP = {
    "int": "int",
    "float": "float",
    "string": "string",
    "bool": "bool",
}

LUA_TYPE_DEFAULTS = {
    "int": 0,
    "float": 0.0,
    "string": '""',
    "bool": "false",
}


def parse_cell_type(raw_type):
    """Parse type string, return (is_list, is_map, key_type, value_type, elem_type)"""
    raw_type = raw_type.strip()

    # map<K,V>
    m = MAP_TYPE_PATTERN.match(raw_type)
    if m:
        return ("map", m.group("key"), m.group("val"), None)

    # list<T>
    m = LIST_TYPE_PATTERN.match(raw_type)
    if m:
        return ("list", None, None, m.group("elem"))

    # scalar
    if raw_type in CSHARP_TYPE_MAP:
        return ("scalar", None, None, raw_type)

    print(f"[WARN] Unknown type: '{raw_type}', treating as string")
    return ("scalar", None, None, "string")


def parse_cell_value(value, type_info, file_name, sheet_name, field_name, col_index, row_idx):
    """Parse cell string value into proper typed value."""
    kind, key_t, val_t, elem_t = type_info

    if value is None or (isinstance(value, str) and value.strip() == ""):
        return _default_value(kind, key_t, val_t, elem_t)

    s = str(value).strip()

    if kind == "scalar":
        return _parse_scalar(s, elem_t, file_name, sheet_name, field_name, col_index, row_idx)

    if kind == "list":
        return _parse_list(s, elem_t, file_name, sheet_name, field_name, col_index, row_idx)

    if kind == "map":
        return _parse_map(s, key_t, val_t, file_name, sheet_name, field_name, col_index, row_idx)

    return s


def _default_value(kind, key_t, val_t, elem_t):
    if kind == "scalar":
        if elem_t == "int":
            return 0
        elif elem_t == "float":
            return 0.0
        elif elem_t == "bool":
            return False
        return ""
    elif kind == "list":
        return []
    elif kind == "map":
        return {}
    return ""


def _col_letter(index):
    """Convert 0-based column index to Excel column letter (A, B, ... Z, AA, AB...)."""
    letter = ""
    i = index
    while i >= 0:
        letter = chr(ord('A') + (i % 26)) + letter
        i = i // 26 - 1
    return letter


def _parse_scalar(s, t, file_name, sheet_name, field_name, col_index, row_idx):
    col_letter = _col_letter(col_index)
    if t == "int":
        try:
            return int(s)
        except ValueError:
            raise ValueError(
                "[TYPE ERROR] File '%s', Sheet '%s', Cell %s%d, Field '%s': expected int, got '%s'"
                % (file_name, sheet_name, col_letter, row_idx, field_name, s)
            )
    elif t == "float":
        try:
            return float(s)
        except ValueError:
            raise ValueError(
                "[TYPE ERROR] File '%s', Sheet '%s', Cell %s%d, Field '%s': expected float, got '%s'"
                % (file_name, sheet_name, col_letter, row_idx, field_name, s)
            )
    elif t == "bool":
        return s.lower() in ("true", "1", "yes")
    return s


def _parse_list(s, elem_t, file_name, sheet_name, field_name, col_index, row_idx):
    items = _split_csv(s)
    result = []
    for item in items:
        item = item.strip()
        if item == "":
            continue
        result.append(_parse_scalar(item, elem_t, file_name, sheet_name, field_name, col_index, row_idx))
    return result


def _parse_map(s, key_t, val_t, file_name, sheet_name, field_name, col_index, row_idx):
    pairs = _split_csv(s)
    result = OrderedDict()
    for pair in pairs:
        pair = pair.strip()
        if pair == "":
            continue
        colon_idx = _find_first_colon(pair)
        if colon_idx < 0:
            col_letter = _col_letter(col_index)
            raise ValueError(
                "[TYPE ERROR] File '%s', Sheet '%s', Cell %s%d, Field '%s': invalid map entry '%s' (missing ':')"
                % (file_name, sheet_name, col_letter, row_idx, field_name, pair)
            )
        k = pair[:colon_idx].strip()
        v = pair[colon_idx + 1:].strip()
        key = _parse_scalar(k, key_t, file_name, sheet_name, field_name, col_index, row_idx)
        val = _parse_scalar(v, val_t, file_name, sheet_name, field_name, col_index, row_idx)
        result[str(key)] = val
    return result


def _split_csv(s):
    """Split by comma, respecting double-quoted strings."""
    result = []
    current = []
    in_quote = False
    for ch in s:
        if ch == '"':
            in_quote = not in_quote
        elif ch == "," and not in_quote:
            result.append("".join(current))
            current = []
        else:
            current.append(ch)
    result.append("".join(current))
    return result


def _find_first_colon(s):
    """Find first colon not inside double quotes."""
    in_quote = False
    for i, ch in enumerate(s):
        if ch == '"':
            in_quote = not in_quote
        elif ch == ":" and not in_quote:
            return i
    return -1


# ============================================================
# Excel Reader
# ============================================================

class TableDef:
    """Represents one config table parsed from Excel."""

    def __init__(self, name):
        self.name = name
        self.fields = []  # [(field_name, type_info), ...]
        self.rows = []    # [OrderedDict, ...]

    def add_field(self, field_name, raw_type):
        type_info = parse_cell_type(raw_type)
        self.fields.append((field_name, type_info))

    def add_row(self, values, file_name, row_idx):
        row = OrderedDict()
        for i, (field_name, type_info) in enumerate(self.fields):
            val = values[i] if i < len(values) else ""
            row[field_name] = parse_cell_value(val, type_info, file_name, self.name, field_name, i, row_idx)
        self.rows.append(row)

    def get_key_field(self):
        """Return first field name (convention: first column is key/id)."""
        if self.fields:
            return self.fields[0][0]
        return None


def read_excel(filepath):
    """Read an Excel file and return a list of TableDef (one per sheet)."""
    file_name = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            print(f"[SKIP] Sheet '{sheet_name}' has < 4 rows (need field+desc+type+data)")
            continue

        # Row 0: Chinese description, ignored.
        # Row 1: field names (English). Skip column if empty.
        field_names = [str(c) if c is not None else "" for c in rows[1]]
        # Row 2: field types.
        field_types = [str(c) if c is not None else "" for c in rows[2]]

        # Filter: only columns with non-empty field name AND non-empty type
        valid_indices = []
        valid_names = []
        valid_types = []
        for i, (name, typ) in enumerate(zip(field_names, field_types)):
            if name.strip() and typ.strip():
                valid_indices.append(i)
                valid_names.append(name.strip())
                valid_types.append(typ.strip())

        if not valid_names:
            print(f"[SKIP] Sheet '{sheet_name}' has no valid columns")
            continue

        table = TableDef(sheet_name)
        for name, typ in zip(valid_names, valid_types):
            table.add_field(name, typ)

        # Data starts from row 3 (0-indexed: row 3 = 4th row in Excel, 1-indexed: row 4)
        for row_idx, row in enumerate(rows[3:], start=4):
            vals = [row[i] if i < len(row) else None for i in valid_indices]
            # Skip completely empty rows
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
                continue
            table.add_row(vals, file_name, row_idx)

        if table.rows:
            tables.append(table)
            print(f"  [OK] Sheet '{sheet_name}': {len(table.fields)} fields, {len(table.rows)} rows")
        else:
            print(f"[SKIP] Sheet '{sheet_name}' has no data rows")

    wb.close()
    return tables


# ============================================================
# JSON Export
# ============================================================

def export_json(table, output_dir):
    """Export table to JSON file."""
    data = {"items": []}
    for row in table.rows:
        data["items"].append(row)

    filepath = os.path.join(output_dir, f"{table.name}.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  [JSON] -> {filepath}")
    return filepath


# ============================================================
# Lua Export
# ============================================================

def export_lua(table, output_dir):
    """Export table to Lua file. Uses first column as key."""
    key_field = table.get_key_field()
    if not key_field:
        return None

    lines = []
    lines.append("-- Auto-generated from Excel: %s" % table.name)
    lines.append("return {")

    for row in table.rows:
        key_val = row.get(key_field, 0)
        lines.append("  [%s] = {" % _lua_repr(key_val))
        for field_name, type_info in table.fields:
            if field_name == key_field:
                continue
            val = row.get(field_name, _default_value(*type_info))
            lines.append("    %s = %s," % (field_name, _lua_repr(val)))
        lines.append("  },")

    lines.append("}")
    content = "\n".join(lines)

    filepath = os.path.join(output_dir, f"{table.name}.lua")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  [LUA] -> {filepath}")
    return filepath


def _lua_repr(val):
    """Convert Python value to Lua literal string."""
    if val is None:
        return "nil"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, str):
        # Escape
        escaped = val.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")
        return '"%s"' % escaped
    if isinstance(val, (list, tuple)):
        items = ", ".join(_lua_repr(v) for v in val)
        return "{%s}" % items
    if isinstance(val, dict):
        items = ", ".join("[%s] = %s" % (_lua_repr(k), _lua_repr(v)) for k, v in val.items())
        return "{%s}" % items
    return '"%s"' % str(val)


# ============================================================
# C# Class Export
# ============================================================

def export_csharp(table, output_dir, namespace):
    """Export table as C# data class."""
    class_name = _pascal_case(table.name)
    lines = []
    lines.append("// Auto-generated from Excel: %s" % table.name)
    lines.append("using System;")
    lines.append("using System.Collections.Generic;")
    lines.append("")
    if namespace:
        lines.append("namespace %s" % namespace)
        lines.append("{")

    lines.append("    [Serializable]")
    lines.append("    public class %s" % class_name)
    lines.append("    {")

    for field_name, type_info in table.fields:
        cs_type = _csharp_type_str(type_info)
        lines.append("        public %s %s;" % (cs_type, field_name))

    lines.append("    }")

    if namespace:
        lines.append("}")

    content = "\n".join(lines)

    filepath = os.path.join(output_dir, "%s.cs" % class_name)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  [CS]  -> {filepath}")
    return filepath


def _pascal_case(name):
    """Convert snake_case or camelCase to PascalCase."""
    # Handle camelCase: insert underscore before uppercase letters
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', name)
    parts = s.replace('-', '_').split('_')
    return ''.join(p.capitalize() for p in parts if p)


def _csharp_type_str(type_info):
    """Convert type_info to C# type string."""
    kind, key_t, val_t, elem_t = type_info
    if kind == "scalar":
        return CSHARP_TYPE_MAP.get(elem_t, "string")
    elif kind == "list":
        inner = CSHARP_TYPE_MAP.get(elem_t, "string")
        return "List<%s>" % inner
    elif kind == "map":
        k = CSHARP_TYPE_MAP.get(key_t, "string")
        v = CSHARP_TYPE_MAP.get(val_t, "string")
        return "Dictionary<%s, %s>" % (k, v)
    return "string"


# ============================================================
# Main Export Logic
# ============================================================

def export_all():
    """Export all Excel files in the Excel directory."""
    print("=" * 60)
    print("Config Table Export Tool")
    print("=" * 60)
    print()

    # Ensure output dirs exist
    for d in [CLIENT_JSON_DIR, CLIENT_LUA_DIR, CLIENT_CS_DIR, SERVER_CS_DIR]:
        os.makedirs(d, exist_ok=True)

    # Find all .xlsx files
    excel_files = [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx") and not f.startswith("~$")]
    if not excel_files:
        print("[WARN] No .xlsx files found in: %s" % EXCEL_DIR)
        print("       Place your Excel files there and run again.")
        return

    total_tables = 0

    for excel_file in sorted(excel_files):
        filepath = os.path.join(EXCEL_DIR, excel_file)
        print("[FILE] %s" % excel_file)

        try:
            tables = read_excel(filepath)
        except ValueError as e:
            print("  [ERROR] %s" % e)
            print("  -> Skipping this file due to type error, fix and re-run.")
            continue

        if not tables:
            print("  -> No valid sheets found, skipping")
            continue

        for table in tables:
            export_json(table, CLIENT_JSON_DIR)
            export_lua(table, CLIENT_LUA_DIR)
            export_csharp(table, CLIENT_CS_DIR, None)
            export_csharp(table, SERVER_CS_DIR, "Config")
            total_tables += 1
            print()

    print("=" * 60)
    print("Done! %d config tables exported." % total_tables)
    print("  JSON -> %s" % CLIENT_JSON_DIR)
    print("  LUA  -> %s" % CLIENT_LUA_DIR)
    print("  C#   -> %s" % CLIENT_CS_DIR)
    print("  C#   -> %s" % SERVER_CS_DIR)
    print("=" * 60)


if __name__ == "__main__":
    export_all()
